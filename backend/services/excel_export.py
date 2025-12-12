from datetime import datetime
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend.db import models


EXPORT_DIR = Path(__file__).resolve().parent.parent / "data" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def fill_workbook_from_assignments(template_path: Path, template_data: Dict, db: Session, workshop_id: int) -> Path:
    workbook = load_workbook(template_path)
    assignments = (
        db.query(models.Assignment)
        .filter(models.Assignment.workshop_id == workshop_id)
        .all()
    )
    assignment_map = {}
    for assignment in assignments:
        assignment_map[(assignment.activity_id, assignment.role_id)] = assignment.raci_value

    # build helper maps
    role_lookup = {role.id: role for role in db.query(models.Role).filter(models.Role.workshop_id == workshop_id)}
    activity_lookup = {activity.id: activity for activity in db.query(models.Activity).filter(models.Activity.workshop_id == workshop_id)}

    for activity_blob in template_data.get("activities", []):
        domain_sheet = activity_blob["domain"]
        if domain_sheet not in workbook.sheetnames:
            continue
        sheet = workbook[domain_sheet]
        cell_map = activity_blob.get("cell_map", {})
        activity_match = next((a for a in activity_lookup.values() if a.activity_text == activity_blob["activity_text"]), None)
        if not activity_match:
            continue
        for role_name, coords in cell_map.items():
            role_obj = next((r for r in role_lookup.values() if r.role_name == role_name), None)
            if not role_obj:
                continue
            value = assignment_map.get((activity_match.id, role_obj.id))
            if value:
                row_idx, col_idx = coords
                sheet[f"{get_column_letter(col_idx)}{row_idx}"] = value

    outputs_sheet = workbook.create_sheet("Outputs")
    outputs_sheet.append(["Exported", datetime.utcnow().isoformat()])
    outputs_sheet.append(["Workshop ID", workshop_id])
    outputs_sheet.append(["Notes", "Filled from OT RACI Workshop Wizard"])

    export_path = EXPORT_DIR / f"workshop_{workshop_id}_filled.xlsx"
    workbook.save(export_path)
    return export_path


def export_actions_csv(db: Session, workshop_id: int) -> Path:
    path = EXPORT_DIR / f"workshop_{workshop_id}_actions.csv"
    import csv

    actions = db.query(models.Issue).filter(models.Issue.workshop_id == workshop_id).all()
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Issue", "Severity", "Status", "Activity"])
        for issue in actions:
            writer.writerow([issue.issue_type, issue.severity, issue.status, issue.activity_id])
    return path
