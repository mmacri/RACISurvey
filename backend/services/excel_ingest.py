import hashlib
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


class ParsedTemplate:
    def __init__(self, template_path: Path):
        self.template_path = template_path
        self.workbook = load_workbook(template_path, data_only=True)
        self.domains: List[Dict] = []
        self.roles: List[Dict] = []
        self.activities: List[Dict] = []
        self.instructions: Dict[str, str] = {}
        self.lists: Dict[str, List[str]] = {}

    def _hash_file(self) -> str:
        data = self.template_path.read_bytes()
        return hashlib.sha256(data).hexdigest()

    def _detect_role_row(self, sheet) -> Tuple[int, List[str]]:
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = list(row)
            if len(cells) < 2:
                continue
            header_candidates = [c for c in cells[1:] if c]
            if len(header_candidates) >= 1:
                return idx, [str(c).strip() for c in cells[1:] if c]
        return 1, []

    def parse_matrix_sheet(self, sheet_name: str):
        sheet = self.workbook[sheet_name]
        role_row_index, role_names = self._detect_role_row(sheet)
        role_map: Dict[int, str] = {i + 1: name for i, name in enumerate(role_names)}
        self.domains.append(
            {
                "sheet_name": sheet_name,
                "display_name": sheet_name,
                "order_index": len(self.domains),
            }
        )

        # register roles globally
        for col_index, name in role_map.items():
            self.roles.append(
                {
                    "role_name": name,
                    "role_key": f"{sheet_name}:{name}",
                    "order_index": col_index,
                    "domain": sheet_name,
                    "column_index": col_index + 1,
                }
            )

        current_section = None
        for r_idx, row in enumerate(sheet.iter_rows(min_row=role_row_index + 1, values_only=True), start=role_row_index + 1):
            first_cell = row[0]
            remaining = row[1:]
            if not any([first_cell, *remaining]):
                continue
            # section header if other cells empty
            if first_cell and not any(remaining):
                current_section = str(first_cell)
                continue
            if not first_cell:
                continue
            activity_text = str(first_cell)
            assignments = {}
            for c_idx, value in enumerate(remaining, start=1):
                if value:
                    assignments[role_map.get(c_idx)] = str(value).strip()
            self.activities.append(
                {
                    "domain": sheet_name,
                    "activity_text": activity_text,
                    "section_text": current_section,
                    "order_index": len(self.activities),
                    "cell_map": {role_map.get(c_idx): (r_idx, c_idx + 1) for c_idx in role_map.keys()},
                    "initial_assignments": assignments,
                }
            )

    def parse_instructions_sheet(self):
        for name in self.workbook.sheetnames:
            if name.lower().startswith("instruction"):
                content_lines = []
                for row in self.workbook[name].iter_rows(values_only=True):
                    line = " ".join([str(c) for c in row if c])
                    if line:
                        content_lines.append(line)
                self.instructions[name] = "\n".join(content_lines)

    def parse_lists_sheet(self):
        for name in self.workbook.sheetnames:
            if name.lower().startswith("list"):
                values: List[str] = []
                for row in self.workbook[name].iter_rows(values_only=True):
                    if row and row[0]:
                        values.append(str(row[0]))
                self.lists[name] = values

    def parse(self):
        for name in self.workbook.sheetnames:
            if name.lower().endswith("raci"):
                self.parse_matrix_sheet(name)
        self.parse_instructions_sheet()
        self.parse_lists_sheet()

        return {
            "file_hash": self._hash_file(),
            "domains": self.domains,
            "roles": self.roles,
            "activities": self.activities,
            "instructions": self.instructions,
            "lists": self.lists,
        }


def parse_template(path: Path) -> Dict:
    parser = ParsedTemplate(path)
    return parser.parse()
