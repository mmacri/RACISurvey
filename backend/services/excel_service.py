from io import BytesIO
from typing import List
from openpyxl import load_workbook, Workbook
from ..schemas import TemplateIn


def parse_workbook(file_bytes: bytes) -> TemplateIn:
  wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
  sections: List[str] = []
  roles: List[str] = []
  activities = []
  for name in wb.sheetnames:
    if 'raci' in name.lower() and name.lower() != 'raci definitions':
      sheet = wb[name]
      current_section = name
      # gather roles from first row
      for cell in sheet[1][1:]:
        if cell.value:
          roles.append(str(cell.value))
      for row in sheet.iter_rows(min_row=2):
        label = row[0].value
        if not label:
          continue
        row_roles = [c.value for c in row[1:] if c.value]
        if not row_roles and len(str(label)) < 60:
          current_section = str(label)
          if current_section not in sections:
            sections.append(current_section)
          continue
        activities.append({
          'id': f"{name}:{current_section}:{label}",
          'section': current_section,
          'sheet': name,
          'text': str(label)
        })
  return TemplateIn(name=wb.properties.title or 'Imported Template', source_filename='upload.xlsx', sections=sections or list(dict.fromkeys([a['section'] for a in activities])), roles=list(dict.fromkeys(roles)), activities=activities)


def fill_workbook(template: TemplateIn, responses: List[dict]) -> bytes:
  wb = Workbook()
  ws = wb.active
  ws.title = 'Filled RACI'
  ws.append(['Activity', 'Accountable', 'Responsible', 'Consulted', 'Informed', 'Confidence', 'Status', 'Notes'])
  for res in responses:
    ws.append([
      res.get('activity_id'), res.get('accountable_role'), ', '.join(res.get('responsible_roles', [])), ', '.join(res.get('consulted_roles', [])), ', '.join(res.get('informed_roles', [])), res.get('confidence'), res.get('status'), res.get('notes', '')
    ])
  stream = BytesIO()
  wb.save(stream)
  return stream.getvalue()
