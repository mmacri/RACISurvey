import os
import base64
import sys
import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from openpyxl import Workbook

# configure test database before importing app
os.environ["RACI_DATABASE_URL"] = "sqlite://"

from backend.main import app  # noqa: E402
from backend.db.database import Base, engine  # noqa: E402


@pytest.fixture(autouse=True)
def reset_db():
    # ensure clean database
    assert "sqlite://" in str(engine.url)
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    yield


def build_sample_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "APPLICATIONS RACI"
    ws["A1"] = "Activity"
    ws["B1"] = "CIO"
    ws["C1"] = "CISO"
    ws["A2"] = "Select OT vendor"
    ws["A3"] = "Deploy patch"
    ws["B2"] = "R"
    ws["C2"] = "A"
    ws["B3"] = "R"
    ws["C3"] = "I"

    inst = wb.create_sheet("Instructions")
    inst["A1"] = "Rules of engagement"

    path = tmp_path / "template.xlsx"
    wb.save(path)
    return path


def test_workflow_from_upload_to_export(tmp_path):
    workbook_path = build_sample_workbook(tmp_path)
    client = TestClient(app)

    with open(workbook_path, "rb") as fh:
        content_b64 = base64.b64encode(fh.read()).decode()
    response = client.post(
        "/api/templates/upload",
        json={"file_name": workbook_path.name, "content": content_b64},
    )
    assert response.status_code == 200
    template_id = response.json()["template"]["id"]
    assert response.json()["domains"]
    assert response.json()["roles"]

    workshop = client.post(
        "/api/workshops",
        json={"template_id": template_id, "org_name": "Contoso", "workshop_name": "OT RACI – Current State"},
    ).json()

    domains = client.get(f"/api/workshops/{workshop['id']}/domains").json()
    roles = client.get(f"/api/workshops/{workshop['id']}/domains/{domains[0]['id']}/roles").json()
    activities = client.get(f"/api/workshops/{workshop['id']}/domains/{domains[0]['id']}/activities").json()
    assert len(activities) == 2

    # send incomplete assignments to trigger issues
    payload = {
        "domain_id": domains[0]["id"],
        "assignments": [
            {"activity_id": activities[0]["id"], "role_id": roles[0]["id"], "raci_value": "R"},
        ],
    }
    client.put(f"/api/workshops/{workshop['id']}/assignments/bulk", json=payload)
    validation = client.post(f"/api/workshops/{workshop['id']}/validate").json()
    assert validation["created"]

    # resolve with an Accountable and Inform
    payload["assignments"].append({"activity_id": activities[0]["id"], "role_id": roles[1]["id"], "raci_value": "A"})
    payload["assignments"].append({"activity_id": activities[0]["id"], "role_id": roles[1]["id"], "raci_value": "I"})
    client.put(f"/api/workshops/{workshop['id']}/assignments/bulk", json=payload)
    progress = client.get(f"/api/workshops/{workshop['id']}/progress").json()
    assert progress["complete"] >= 1

    excel = client.post(f"/api/workshops/{workshop['id']}/export/excel")
    assert excel.status_code == 200

    pdf = client.post(f"/api/workshops/{workshop['id']}/export/pdf")
    assert pdf.status_code == 200

    pptx = client.post(f"/api/workshops/{workshop['id']}/export/pptx")
    assert pptx.status_code == 200
